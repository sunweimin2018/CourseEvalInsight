"""
Excel generation for Module 5 — Course Objective Achievement Calculation.

Produces an .xlsx matching the structure of the reference template
docs/templates/质量监测及达成度(新版).xlsx, with dynamic column layout
that adapts to any number of objectives and evaluation items.
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from course_eval.apps.excel.data_handler import get_effective_data
from .report_engine import (
    _parse_syllabus,
    _extract_score_weights,
    _parse_wide_eval_table,
    _convert_wide_to_long_mapping,
    _parse_objective_item_mapping,
    _match_score_weights,
    _infer_objective_mapping,
    _compute_objective_achievement,
    _parse_eval_item_totals,
    _find_name_column_index,
    _is_score_column,
)


# ── Styling ─────────────────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FONT = Font(name='宋体', size=10, bold=True)
NORMAL_FONT = Font(name='宋体', size=10)
TITLE_FONT = Font(name='宋体', size=14, bold=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _style(cell, font=NORMAL_FONT, alignment=CENTER):
    cell.font = font
    cell.alignment = alignment
    cell.border = THIN_BORDER


def _sc(ws, row, col, value, font=NORMAL_FONT, alignment=CENTER):
    """Set value + style on a cell, return the cell."""
    c = ws.cell(row=row, column=col, value=value)
    _style(c, font, alignment)
    return c


# ── Helpers ─────────────────────────────────────────────────────────────────

def _parse_score(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_numeric(val, default=0.0):
    """Parse a value that might be a string with %, or a plain number."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace('%', '').replace('％', '').strip()
        return float(s) if s else default
    except (ValueError, TypeError):
        return default


def _build_objective_mapping(evaluation_standards, objectives, score_columns, headers, rows):
    item_weights_map = _extract_score_weights(evaluation_standards) if evaluation_standards else None
    wide_data = _parse_wide_eval_table(evaluation_standards) if evaluation_standards else None
    if wide_data:
        objective_mapping = _convert_wide_to_long_mapping(wide_data)
    else:
        objective_mapping = (
            _parse_objective_item_mapping(evaluation_standards) if evaluation_standards else None
        )
    if not objective_mapping:
        column_weights = _match_score_weights(score_columns, item_weights_map) if item_weights_map else {}
        objective_mapping = _infer_objective_mapping(
            objectives, score_columns, column_weights, evaluation_standards, headers, rows)
    return objective_mapping


def _find_name_col(headers):
    for h in headers:
        if '姓名' in str(h).replace(' ', '').strip():
            return h
    return headers[0] if headers else None


def _find_total_col(headers):
    for h in headers:
        norm = str(h).replace(' ', '').strip()
        if norm in ('总评', '总评成绩', '总分', '总成绩'):
            return h
    for h in headers:
        norm = str(h).replace(' ', '').strip()
        if '总评' in norm or '总分' in norm or '总成绩' in norm:
            return h
    return None


def _compute_col_max(rows, col_name):
    """Compute the maximum observed value for a column."""
    mx = 0
    for row in rows:
        v = _parse_score(row.get(col_name))
        if v is not None and v > mx:
            mx = v
    return mx


def _effective_divisor(item_name, item_totals, col_max):
    """Determine the correct divisor for contribution calculation."""
    it = item_totals.get(item_name) if item_totals else None
    if it is not None:
        it = _parse_numeric(it)
    if it is None or it == 0:
        return max(col_max, 1)
    if col_max > it * 1.5:
        return max(col_max, 100)
    return max(it, col_max)


# ── Main entry point ────────────────────────────────────────────────────────

def generate_achievement_excel(report):
    """Generate the achievement calculation Excel.

    Returns BytesIO buffer, or None if data is insufficient.
    """
    if not report.grades_file or not report.user_id:
        return None

    try:
        grades_data = get_effective_data(report.grades_file, report.user_id)
    except Exception:
        return None

    headers = grades_data['headers']
    rows = grades_data['rows']

    # Parse syllabus
    syllabus_fields = _parse_syllabus(report.syllabus_file)
    evaluation_standards = syllabus_fields.get('evaluation_standards')
    objectives_str = report.report_data.get('module_2_objectives', '')

    # Identify score columns
    name_col_index = _find_name_column_index(headers)
    score_columns = []
    for i, col in enumerate(headers):
        col_values = [row.get(col) for row in rows if col in row]
        if _is_score_column(col, col_values, i, name_col_index):
            score_columns.append(col)
    if not score_columns:
        return None

    # Build objective mapping
    objective_mapping = _build_objective_mapping(
        evaluation_standards, objectives_str, score_columns, headers, rows)
    if not objective_mapping:
        return None

    # Compute achievement
    item_totals = _parse_eval_item_totals(evaluation_standards)
    ach_data = _compute_objective_achievement(
        headers, rows, objective_mapping, score_columns, item_totals)

    per_student = ach_data['per_student']
    objective_items = ach_data['objective_items']
    objective_names = ach_data['objective_names']
    item_to_col = ach_data['item_to_col']
    col_avgs = ach_data['col_avgs']
    obj_rates = ach_data['obj_rates']

    if not objective_names:
        return None

    name_col = _find_name_col(headers)
    total_score_col = _find_total_col(headers)

    # ── Compute effective divisors per item ──────────────────────────────
    col_max_map = {}
    for item_name, col_name in item_to_col.items():
        col_max_map[item_name] = _compute_col_max(rows, col_name)

    # ── Build ordered item list and per-item metadata ────────────────────
    item_order = []
    seen = set()
    for m in objective_mapping:
        if m['item'] not in seen:
            seen.add(m['item'])
            item_order.append(m['item'])
    for obj_name in objective_names:
        for item_name, _ts, _wp in objective_items.get(obj_name, []):
            if item_name not in seen:
                seen.add(item_name)
                item_order.append(item_name)

    num_items = len(item_order)

    # ── Column layout ────────────────────────────────────────────────────
    # A=序号  B=姓名  C..=items  总分  [per-obj sub-items] [达成度]
    ITEM_START = 3  # column C
    TOTAL_COL = ITEM_START + num_items

    obj_cols = []  # (name, start, end, ach_col)
    cur = TOTAL_COL + 1
    for obj_name in objective_names:
        oi = objective_items.get(obj_name, [])
        s = cur
        e = cur + len(oi) - 1
        a = cur + len(oi)
        obj_cols.append((obj_name, s, e, a))
        cur = a + 1
    LAST_COL = cur - 1

    # ── Build Workbook ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = '课程达成度计算'

    # --- Row 1: Title ---
    sem = report.semester.name if report.semester else ''
    cn = report.course.name if report.course else ''
    title = f'{sem}《{cn}》课程达成度计算'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)
    _sc(ws, 1, 1, title, font=TITLE_FONT)

    # --- Row 2: Objective group headers ---
    for oname, sc, ec, ac in obj_cols:
        if sc <= ac:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ac)
        _sc(ws, 2, sc, oname, font=HEADER_FONT)

    # --- Row 3: Column sub-headers ---
    _sc(ws, 3, 1, '序号', font=HEADER_FONT)
    _sc(ws, 3, 2, '姓名', font=HEADER_FONT)
    for idx, iname in enumerate(item_order):
        # Find weight_pct for this item from any mapping entry
        wp = 0.0
        for m in objective_mapping:
            if m['item'] == iname:
                wp = _parse_numeric(m.get('weight_pct', 0))
                break
        if wp == 0:
            for oname in objective_names:
                for inm, _ts, wgt in objective_items.get(oname, []):
                    if inm == iname:
                        wp = _parse_numeric(wgt)
                        break
        # weight_pct from parser is already a percentage (e.g. 4 = 4%)
        # but may also be a decimal (e.g. 0.04 = 4%) from other sources
        pct = wp * 100 if wp < 1 else wp
        label = f'{iname}{round(pct)}%' if pct else iname
        _sc(ws, 3, ITEM_START + idx, label, font=HEADER_FONT)
    _sc(ws, 3, TOTAL_COL, '总分\n(100)', font=HEADER_FONT)

    for oname, sc, ec, ac in obj_cols:
        oi = objective_items.get(oname, [])
        for j, (iname, _ts, wp) in enumerate(oi):
            _sc(ws, 3, sc + j, f'{iname}{round(_parse_numeric(wp), 4)}', font=HEADER_FONT)
        _sc(ws, 3, ac, f'{oname}达成度', font=HEADER_FONT)

    # --- Row 4: 目标分值 ---
    _sc(ws, 4, 2, '目标分值', font=HEADER_FONT)
    for idx, iname in enumerate(item_order):
        it = item_totals.get(iname) if item_totals else None
        if it is not None:
            it = _parse_numeric(it)
        if it is None:
            it = col_max_map.get(iname, '')
        _sc(ws, 4, ITEM_START + idx, it)

    total_target = 0
    for n in item_order:
        if item_totals and item_totals.get(n):
            total_target += _parse_numeric(item_totals.get(n))
        else:
            total_target += col_max_map.get(n, 0)
    _sc(ws, 4, TOTAL_COL, total_target if total_target else 100)

    for oname, sc, ec, ac in obj_cols:
        oi = objective_items.get(oname, [])
        obj_total = 0
        for j, (iname, ts, _wp) in enumerate(oi):
            ts_val = _parse_numeric(ts)
            _sc(ws, 4, sc + j, ts_val)
            obj_total += ts_val
        _sc(ws, 4, ac, obj_total)

    # --- Rows 6..N: Student data ---
    DATA_START = 6

    # Build name→grade_row lookup
    name_to_row = {}
    for row in rows:
        n = str(row.get(name_col, '')).strip() if name_col else ''
        if n:
            name_to_row[n] = row

    for si, ps in enumerate(per_student):
        rn = DATA_START + si
        sname = ps['name']
        gr = name_to_row.get(sname, {})

        _sc(ws, rn, 1, si + 1)
        _sc(ws, rn, 2, sname, alignment=LEFT)

        # Raw item scores
        for idx, iname in enumerate(item_order):
            col_name = item_to_col.get(iname)
            if col_name:
                raw = _parse_score(gr.get(col_name))
                if raw is not None:
                    _sc(ws, rn, ITEM_START + idx, raw)

        # Total score
        if total_score_col:
            ts_raw = _parse_score(gr.get(total_score_col))
            if ts_raw is not None:
                _sc(ws, rn, TOTAL_COL, ts_raw)

        # Per-objective weighted contributions and achievement
        for oname, sc, ec, ac in obj_cols:
            oi = objective_items.get(oname, [])
            contributions = []
            for j, (iname, ts, _wp) in enumerate(oi):
                col_name = item_to_col.get(iname)
                if not col_name:
                    continue
                raw = _parse_score(gr.get(col_name))
                if raw is None:
                    continue
                ts_val = _parse_numeric(ts)
                divisor = _effective_divisor(iname, item_totals, col_max_map.get(iname, 0))
                if divisor > 0:
                    contrib = raw * ts_val / divisor
                    contributions.append(contrib)
                    _sc(ws, rn, sc + j, round(contrib, 4))

            obj_target_total = sum(_parse_numeric(t) for _, t, _ in oi)
            if obj_target_total > 0:
                ach_rate = sum(contributions) / obj_target_total
                _sc(ws, rn, ac, round(ach_rate, 4))

    LAST_DATA = DATA_START + len(per_student) - 1

    # --- 班级平均达成度 row ---
    avg_row = LAST_DATA + 1
    _sc(ws, avg_row, 1, '班级平均达成度', font=HEADER_FONT)
    ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=TOTAL_COL - 1)

    for _oname, sc, ec, ac in obj_cols:
        for c in range(sc, ec + 1):
            cl = get_column_letter(c)
            if LAST_DATA >= DATA_START:
                ws.cell(row=avg_row, column=c).value = (
                    f'=AVERAGE({cl}{DATA_START}:{cl}{LAST_DATA})')
                _style(ws.cell(row=avg_row, column=c))
        al = get_column_letter(ac)
        if LAST_DATA >= DATA_START:
            ws.cell(row=avg_row, column=ac).value = (
                f'=AVERAGE({al}{DATA_START}:{al}{LAST_DATA})')
            _style(ws.cell(row=avg_row, column=ac))

    # --- 平均值 row ---
    mean_row = avg_row + 1
    _sc(ws, mean_row, 2, '平均值', font=HEADER_FONT)
    for idx in range(num_items):
        cl = get_column_letter(ITEM_START + idx)
        if LAST_DATA >= DATA_START:
            ws.cell(row=mean_row, column=ITEM_START + idx).value = (
                f'=AVERAGE({cl}{DATA_START}:{cl}{LAST_DATA})')
            _style(ws.cell(row=mean_row, column=ITEM_START + idx))
    tl = get_column_letter(TOTAL_COL)
    if LAST_DATA >= DATA_START:
        ws.cell(row=mean_row, column=TOTAL_COL).value = (
            f'=AVERAGE({tl}{DATA_START}:{tl}{LAST_DATA})')
        _style(ws.cell(row=mean_row, column=TOTAL_COL))

    # ── Section B: 课程目标达成情况 summary ──────────────────────────────
    AUX_COL = LAST_COL + 3
    AUX_ROW = DATA_START

    aux_hdrs = ['课程目标', '评价内容', '目标分值', '平均得分', '权重系数', '课程目标达成情况']
    for j, h in enumerate(aux_hdrs):
        _sc(ws, AUX_ROW, AUX_COL + j, h, font=HEADER_FONT)

    ar = AUX_ROW + 1
    for oname in objective_names:
        oi = objective_items.get(oname, [])
        first = True
        grp_start = ar
        for iname, ts, wp in oi:
            if first:
                _sc(ws, ar, AUX_COL, oname, font=HEADER_FONT)
                first = False
            _sc(ws, ar, AUX_COL + 1, iname)
            _sc(ws, ar, AUX_COL + 2, _parse_numeric(ts))

            # Average score — try col_avgs first, else AVERAGE formula
            col_name = item_to_col.get(iname)
            if col_name and col_name in col_avgs:
                _sc(ws, ar, AUX_COL + 3, round(col_avgs[col_name], 4))
            else:
                # Find the item column index for an AVERAGE formula
                for idx, inm in enumerate(item_order):
                    if inm == iname:
                        cl = get_column_letter(ITEM_START + idx)
                        if LAST_DATA >= DATA_START:
                            ws.cell(row=ar, column=AUX_COL + 3).value = (
                                f'=AVERAGE({cl}{DATA_START}:{cl}{LAST_DATA})')
                            _style(ws.cell(row=ar, column=AUX_COL + 3))
                        break

            _sc(ws, ar, AUX_COL + 4, round(_parse_numeric(wp) * 100, 2) / 100)  # decimal
            ar += 1

        grp_end = ar - 1
        # Merge achievement rate cell
        ws.merge_cells(start_row=grp_start, start_column=AUX_COL + 5,
                       end_row=grp_end, end_column=AUX_COL + 5)
        ach_val = obj_rates.get(oname)
        if ach_val is not None:
            _sc(ws, grp_start, AUX_COL + 5, f'{round(ach_val, 2)}%')
        else:
            _sc(ws, grp_start, AUX_COL + 5, '')

        # Merge objective name cells
        if grp_end > grp_start:
            ws.merge_cells(start_row=grp_start, start_column=AUX_COL,
                           end_row=grp_end, end_column=AUX_COL)

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    for idx in range(num_items):
        ws.column_dimensions[get_column_letter(ITEM_START + idx)].width = 12
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 10
    for _oname, sc, _ec, ac in obj_cols:
        for c in range(sc, ac + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12
    for j in range(6):
        ws.column_dimensions[get_column_letter(AUX_COL + j)].width = 14

    # ── Row heights ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    for r in range(2, LAST_DATA + 3):
        ws.row_dimensions[r].height = 20

    # ── Freeze panes ──────────────────────────────────────────────────────
    ws.freeze_panes = 'A5'

    # ── Save ──────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
